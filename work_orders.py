"""
work_orders.py - Work order JSON storage + CRUD API.

Features
- JSON persistence at DB_PATH/work_orders.json
- CRUD endpoints for work orders
- Completed work orders become knowledge candidates.
- Admin approval ingests reviewed notes into the RAG knowledge base.
"""

import copy
import json
import logging
import os
import shutil
import tempfile
import uuid
from datetime import datetime
from typing import List, Optional, Tuple

from fastapi import APIRouter, Depends, File, Query, UploadFile
from pydantic import BaseModel

from api_schemas import (
    API_ERROR_RESPONSES,
    KnowledgeReviewErrorResponse,
    KnowledgeReviewResponse,
    WorkOrderArchiveResponse,
    WorkOrderDeleteResponse,
    WorkOrderHistoryResponse,
    WorkOrderImportResponse,
    WorkOrderMutationResponse,
    WorkOrderStatsResponse,
    WorkOrderSuccessResponse,
    WorkOrdersPageResponse,
    WorkOrdersResponse,
)
from audit_history import append_history, field_changes, history_list
from auth import (
    actor_id,
    actor_role,
    can_reference_rag_answer,
    can_update_work_order,
    can_view_work_order,
    can_verify,
    get_actor,
    is_admin,
    resolve_user,
)
from config_values import env_float, env_int
from pagination import InvalidCursor, decode_cursor, encode_cursor, paginate_records
from repositories.postgres_workflow import PostgresWorkOrderRepository
from repositories.rag_answers import RagAnswerRepository
from repositories.runtime import postgres_store_enabled
from services.json_file_store import write_json_atomic
from services.transactions import postgres_transactional
from services import work_order_lifecycle
from services import work_order_issue_sync
from services import work_order_mutations
from services import work_order_knowledge
from services import work_order_operations
from services import work_order_queries
from services import work_order_reporting
from services.work_order_import import (
    XlsxArchiveLimits,
    detect_columns,
    import_workbook_rows,
    validate_xlsx_archive,
)

logger = logging.getLogger("alarm_rag.work_orders")
router = APIRouter()
postgres_work_orders = PostgresWorkOrderRepository()
rag_answers = RagAnswerRepository()

DB_DIR = os.getenv("DB_PATH", "./alarm_db")
WO_FILE = os.path.join(DB_DIR, "work_orders.json")
ARCHIVE_DIR = os.path.join(DB_DIR, "archive")
STATUSES = ["pending", "assigned", "in_progress", "completed", "verified", "cancelled"]
STATUS_TRANSITIONS = {
    "pending": {"assigned", "in_progress", "cancelled"},
    "assigned": {"pending", "in_progress", "cancelled"},
    "in_progress": {"pending", "assigned", "completed", "cancelled"},
    "completed": {"verified"},
    "verified": set(),
    "cancelled": set(),
}
PRIORITIES = ["low", "medium", "high", "critical"]
WORK_ORDER_STALE_UPDATE_MESSAGE = "Work order changed since you loaded it. Reload and retry."
KB_REVIEW_STATUSES = [
    "not_ready",
    "pending_review",
    "needs_revision",
    "rejected",
    "ingested",
    "validation_failed",
]
OPERATOR_WORK_ORDER_PATCH_FIELDS = {"status", "verified_by", "notes", "updated_by", "version"}
MAINTENANCE_WORK_ORDER_PATCH_FIELDS = {
    "status",
    "priority",
    "assigned_to",
    "machine_id",
    "description",
    "resolution",
    "notes",
    "accepted_by",
    "completed_by",
    "root_cause",
    "repair_action",
    "failure_category",
    "llm_correctness",
    "llm_coverage",
    "llm_missing_info",
    "llm_expected_fix",
    "llm_answer_used",
    "updated_by",
    "version",
}


def upload_limit_bytes(env_name: str, default_mb: float) -> int:
    mb = env_float(env_name, default_mb, minimum=0.000001)
    return max(int(mb * 1024 * 1024), 1)


EXCEL_UPLOAD_MAX_BYTES = upload_limit_bytes("ALARM_RAG_EXCEL_UPLOAD_MAX_MB", 10)
XLSX_MAGIC = b"PK\x03\x04"
XLSX_MAX_UNCOMPRESSED_BYTES = upload_limit_bytes("ALARM_RAG_XLSX_UNCOMPRESSED_MAX_MB", 50)
XLSX_MAX_SHARED_STRINGS_BYTES = upload_limit_bytes("ALARM_RAG_XLSX_SHARED_STRINGS_MAX_MB", 10)
XLSX_MAX_ENTRIES = env_int("ALARM_RAG_XLSX_MAX_ENTRIES", 2000, minimum=1)
XLSX_MAX_COMPRESSION_RATIO = env_float("ALARM_RAG_XLSX_MAX_COMPRESSION_RATIO", 100, minimum=1.0)

STATUS_LABELS = {
    "pending": "待處理",
    "assigned": "已指派",
    "in_progress": "處理中",
    "completed": "已完成",
    "verified": "已驗證",
    "cancelled": "已取消",
}

PRIORITY_LABELS = {
    "low": "低",
    "medium": "中",
    "high": "高",
    "critical": "緊急",
}


def _clean_text(value: Optional[str]) -> str:
    return work_order_lifecycle.clean_text(value)


def _is_operator_or_supervisor(user_id: str) -> bool:
    normalized = _clean_text(user_id).lower()
    actor = resolve_user(normalized)
    return can_verify(actor) or normalized.startswith("operator") or normalized.startswith("supervisor")


def _status_transition_error(previous_status: str, next_status: str) -> str:
    return work_order_lifecycle.status_transition_error(previous_status, next_status, STATUS_TRANSITIONS)


def _closure_error(order: dict) -> str:
    return work_order_lifecycle.closure_error(order, _is_operator_or_supervisor)


def _direct_verification_error(order: dict, user_id: str) -> str:
    return work_order_lifecycle.direct_verification_error(order, user_id, _is_operator_or_supervisor)


def _knowledge_candidate_ready(order: dict) -> bool:
    return work_order_lifecycle.knowledge_candidate_ready(order)


def _refresh_knowledge_review_state(order: dict, changed_fields: set[str] | list[str]) -> None:
    work_order_lifecycle.refresh_knowledge_review_state(order, changed_fields)


def _knowledge_comparison_text(order: dict) -> str:
    return work_order_lifecycle.knowledge_comparison_text(order)


def _find_duplicate_knowledge_order(order: dict, orders: List[dict]) -> str:
    return work_order_lifecycle.find_duplicate_knowledge_order(order, orders)


def _request_fields(req: BaseModel) -> set[str]:
    return work_order_lifecycle.request_fields(req)


def _work_order_patch_permission_error(actor: dict, req: BaseModel) -> str:
    return work_order_lifecycle.patch_permission_error(
        actor_role(actor),
        req,
        operator_fields=OPERATOR_WORK_ORDER_PATCH_FIELDS,
        maintenance_fields=MAINTENANCE_WORK_ORDER_PATCH_FIELDS,
    )


def _parse_iso(value: str) -> Optional[datetime]:
    return work_order_reporting.parse_iso(value)

def _now_like(value: datetime) -> datetime:
    return work_order_reporting.now_like(value)

def _recent_day_keys(days: int) -> List[str]:
    return work_order_reporting.recent_day_keys(days)


# ----------------- Persistence helpers -----------------
def _load_orders() -> List[dict]:
    if postgres_store_enabled():
        return _normalize_loaded_orders(postgres_work_orders.load_all())
    if not os.path.exists(WO_FILE):
        return []
    try:
        with open(WO_FILE, "r", encoding="utf-8") as f:
            return _normalize_loaded_orders(json.load(f))
    except (json.JSONDecodeError, OSError, UnicodeError):
        return []


def _normalize_loaded_orders(payload: object) -> List[dict]:
    if not isinstance(payload, list):
        return []
    orders = [order for order in payload if isinstance(order, dict)]
    for order in orders:
        try:
            order["version"] = max(int(order.get("version") or 1), 1)
        except (TypeError, ValueError):
            order["version"] = 1
        order.setdefault("kb_candidate", False)
        order.setdefault("kb_review_status", "not_ready")
        order.setdefault("kb_review_note", "")
        order.setdefault("kb_reviewed_by", "")
        order.setdefault("kb_reviewed_at", "")
        order.setdefault("kb_ingested_at", "")
        order.setdefault("kb_ingest_result", None)
        order.setdefault("kb_duplicate_of", "")
        _refresh_knowledge_review_state(order, [])
    return orders


def _save_orders(orders: List[dict]):
    if postgres_store_enabled():
        postgres_work_orders.save_all(orders)
        return
    write_json_atomic(WO_FILE, orders)


def _find_order(order_id: str) -> Tuple[int, Optional[dict]]:
    if postgres_store_enabled():
        return -1, postgres_work_orders.get_one(order_id)
    orders = _load_orders()
    for i, o in enumerate(orders):
        if o["id"] == order_id:
            return i, o
    return -1, None


def get_order_dict(order_id: str) -> Optional[dict]:
    _, order = _find_order(order_id)
    return order


def _issue_map_by_id() -> dict[str, dict]:
    try:
        from issues import _load_issues
        return {
            str(issue.get("issue_id") or ""): issue
            for issue in _load_issues()
            if issue.get("issue_id")
        }
    except Exception as exc:
        logger.warning("Issue visibility map failed: %s", exc)
        return {}


def _restore_json_issue(snapshot: dict) -> None:
    from issues import _load_issues, _save_issues

    issue_id = str(snapshot.get("issue_id") or "")
    issues = _load_issues()
    for index, issue in enumerate(issues):
        if str(issue.get("issue_id") or "") == issue_id:
            issues[index] = copy.deepcopy(snapshot)
            _save_issues(issues)
            return


def _operation_dependencies() -> work_order_operations.OperationDependencies:
    from issues import get_issue_dict, sync_issue_from_work_order, unlink_issue_from_work_order

    return work_order_operations.OperationDependencies(
        get_one=postgres_work_orders.get_one,
        save_one=postgres_work_orders.save_one,
        load_all=_load_orders,
        save_all=_save_orders,
        get_issue=get_issue_dict,
        sync_issue=sync_issue_from_work_order,
        unlink_issue=unlink_issue_from_work_order,
        restore_issue=_restore_json_issue,
        append_history=_append_order_history,
        calculate_field_changes=field_changes,
        apply_soft_delete=work_order_mutations.apply_soft_delete,
        logger=logger,
    )


def _query_dependencies() -> work_order_queries.QueryDependencies:
    from issues import get_issue_dict

    return work_order_queries.QueryDependencies(
        find_order=_find_order,
        get_issue=get_issue_dict,
        can_view=can_view_work_order,
        history_list=history_list,
        logger=logger,
    )


def _visible_orders(actor: dict) -> List[dict]:
    issues_by_id = _issue_map_by_id()
    return work_order_reporting.filter_visible_orders(
        _load_orders(),
        actor,
        issues_by_id,
        can_view_work_order,
    )


def validate_issue_verification(work_order_id: str, user_id: str) -> str:
    _, order = _find_order(work_order_id)
    if order is None:
        return f"Work order {work_order_id} not found"
    if order.get("status") != "completed":
        return "The linked work order must be completed before verification."
    if not (
        _clean_text(order.get("root_cause"))
        and _clean_text(order.get("repair_action"))
    ):
        return "The linked work order requires root_cause and repair_action before verification."
    if not _is_operator_or_supervisor(user_id):
        return "Verification must be performed by an operator or supervisor."
    return ""


def _load_archived_orders() -> Tuple[List[dict], List[dict]]:
    return work_order_reporting.load_archived_orders(ARCHIVE_DIR, logger)


# ----------------- Models -----------------
class CreateWorkOrder(BaseModel):
    alarm_code: str
    manual: Optional[str] = "808d"
    issue_id: Optional[str] = ""
    machine_id: Optional[str] = ""
    priority: Optional[str] = "medium"
    assigned_to: Optional[str] = ""
    description: Optional[str] = ""
    rag_suggestion: Optional[str] = ""
    rag_answer_id: Optional[str] = ""
    source: Optional[str] = "manual"  # manual | auto | n8n
    created_by: Optional[str] = ""


class UpdateWorkOrder(BaseModel):
    status: Optional[str] = None
    priority: Optional[str] = None
    assigned_to: Optional[str] = None
    machine_id: Optional[str] = None
    description: Optional[str] = None
    resolution: Optional[str] = None
    notes: Optional[str] = None
    accepted_by: Optional[str] = None
    completed_by: Optional[str] = None
    verified_by: Optional[str] = None
    root_cause: Optional[str] = None
    repair_action: Optional[str] = None
    failure_category: Optional[str] = None
    llm_correctness: Optional[str] = None
    llm_coverage: Optional[str] = None
    llm_missing_info: Optional[str] = None
    llm_expected_fix: Optional[str] = None
    llm_answer_used: Optional[bool] = None
    kb_candidate: Optional[bool] = None
    updated_by: Optional[str] = None
    version: Optional[int] = None


class KnowledgeReviewRequest(BaseModel):
    action: str
    note: Optional[str] = ""
    version: Optional[int] = None


def _build_order_dict(
    alarm_code: str,
    manual: str = "808d",
    machine_id: str = "",
    priority: str = "medium",
    description: str = "",
    rag_suggestion: str = "",
    rag_answer_id: str = "",
    source: str = "auto",
    assigned_to: str = "",
    issue_id: str = "",
    created_by: str = "",
) -> dict:
    """Build a new work order without publishing it."""
    now = datetime.now().isoformat()
    initial_status = "assigned" if _clean_text(assigned_to) else "pending"
    order = {
        "id": str(uuid.uuid4())[:8],
        "issue_id": issue_id,
        "alarm_code": alarm_code,
        "manual": manual,
        "machine_id": machine_id,
        "status": initial_status,
        "priority": priority,
        "assigned_to": assigned_to,
        "created_by": created_by,
        "updated_by": created_by,
        "accepted_by": "",
        "completed_by": "",
        "verified_by": "",
        "description": description,
        "resolution": "",
        "notes": "",
        "root_cause": "",
        "repair_action": "",
        "failure_category": "",
        "llm_correctness": "",
        "llm_coverage": "",
        "llm_missing_info": "",
        "llm_expected_fix": "",
        "llm_answer_used": False,
        "kb_candidate": False,
        "kb_review_status": "not_ready",
        "kb_review_note": "",
        "kb_reviewed_by": "",
        "kb_reviewed_at": "",
        "kb_ingested_at": "",
        "kb_ingest_result": None,
        "kb_duplicate_of": "",
        "rag_suggestion": rag_suggestion,
        "rag_answer_id": rag_answer_id,
        "source": source,
        "created_at": now,
        "updated_at": now,
        "completed_at": "",
        "version": 1,
        "work_order_history": [{
            "action": "created",
            "user_id": created_by,
            "fields": ["assigned_to"] if _clean_text(assigned_to) else [],
            "from_status": "",
            "to_status": initial_status,
            "created_at": now,
        }],
    }
    return order


def _persist_new_order(order: dict) -> dict:
    if postgres_store_enabled():
        order = postgres_work_orders.save_one(order)
    else:
        orders = _load_orders()
        orders.insert(0, order)
        _save_orders(orders)
    logger.info("Created work order %s for alarm %s", order["id"], order.get("alarm_code", ""))
    return order


# ----------------- Public helper -----------------
def create_order_dict(
    alarm_code: str,
    manual: str = "808d",
    machine_id: str = "",
    priority: str = "medium",
    description: str = "",
    rag_suggestion: str = "",
    rag_answer_id: str = "",
    source: str = "auto",
    assigned_to: str = "",
    issue_id: str = "",
    created_by: str = "",
) -> dict:
    """Create and persist a new work order. Returns the order dict."""
    order = _build_order_dict(
        alarm_code=alarm_code,
        manual=manual,
        machine_id=machine_id,
        priority=priority,
        description=description,
        rag_suggestion=rag_suggestion,
        rag_answer_id=rag_answer_id,
        source=source,
        assigned_to=assigned_to,
        issue_id=issue_id,
        created_by=created_by,
    )
    return _persist_new_order(order)


def _append_order_history(
    order: dict,
    action: str,
    user_id: str = "",
    fields: Optional[List[str]] = None,
    from_status: str = "",
    to_status: str = "",
    changes: Optional[List[dict]] = None,
) -> None:
    append_history(order, "work_order_history", action, user_id, fields, from_status, to_status, changes)


def sync_work_order_from_issue(issue: dict, user_id: str = "", note: str = "") -> Optional[dict]:
    work_order_id = str(issue.get("work_order_id") or "")
    if not work_order_id:
        return None

    if postgres_store_enabled():
        order = postgres_work_orders.get_one(work_order_id)
        orders = [order] if order is not None else []
    else:
        orders = _load_orders()
    synced_order = None
    for order in orders:
        if order.get("id") != work_order_id:
            continue
        synced_order = work_order_issue_sync.apply_issue_update(
            order,
            issue,
            work_order_id=work_order_id,
            user_id=user_id,
            note=note,
            now=datetime.now().isoformat(),
            validate_verification=validate_issue_verification,
            status_transition_error=_status_transition_error,
            append_history=_append_order_history,
            calculate_field_changes=field_changes,
        )
        break

    if synced_order is not None:
        if postgres_store_enabled():
            synced_order = postgres_work_orders.save_one(synced_order)
        else:
            _save_orders(orders)
    return synced_order


# ----------------- API routes -----------------
@router.post(
    "/work-orders",
    responses={200: {"model": WorkOrderSuccessResponse}, **API_ERROR_RESPONSES},
)
@postgres_transactional
async def api_create_order(req: CreateWorkOrder, actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    if actor_role(actor) not in ("maintenance", "supervisor", "admin"):
        return {"status": "error", "message": "Permission denied"}
    linked_issue = None
    if req.issue_id:
        from issues import get_issue_dict

        linked_issue = get_issue_dict(req.issue_id)
        if linked_issue is None:
            return {"status": "error", "message": "Issue not found"}
        if not can_view_work_order(actor, {"status": "pending", "assigned_to": ""}, linked_issue):
            return {"status": "error", "message": "Permission denied"}
    rag_answer_id = req.rag_answer_id or str((linked_issue or {}).get("rag_answer_id") or "")
    if rag_answer_id:
        answer = rag_answers.get(rag_answer_id)
        if answer is None:
            return {"status": "error", "message": "RAG answer not found"}
        inherited_from_visible_issue = bool(
            linked_issue
            and not req.rag_answer_id
            and str(linked_issue.get("rag_answer_id") or "") == rag_answer_id
        )
        if not inherited_from_visible_issue and not can_reference_rag_answer(actor, answer):
            return {"status": "error", "message": "Permission denied"}
    order = create_order_dict(
        alarm_code=req.alarm_code,
        manual=req.manual or "808d",
        issue_id=req.issue_id or "",
        machine_id=req.machine_id or "",
        priority=req.priority or "medium",
        description=req.description or "",
        rag_suggestion=req.rag_suggestion or "",
        rag_answer_id=rag_answer_id,
        source=req.source or "manual",
        assigned_to=req.assigned_to or "",
        created_by=actor_id(actor),
    )
    return {"status": "ok", "order": order}


@router.get(
    "/work-orders",
    responses={200: {"model": WorkOrdersResponse}, **API_ERROR_RESPONSES},
)
async def api_list_orders(status: Optional[str] = None, actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    orders = _visible_orders(actor)
    if status:
        orders = [o for o in orders if o["status"] == status]
    return {"total": len(orders), "orders": orders}


@router.get(
    "/work-orders/page",
    responses={200: {"model": WorkOrdersPageResponse}, **API_ERROR_RESPONSES},
)
async def api_page_orders(
    limit: int = Query(default=50, ge=1, le=200),
    cursor: str = "",
    status: Optional[str] = None,
    actor: dict = Depends(get_actor),
):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    try:
        decoded = decode_cursor(cursor)
    except InvalidCursor as exc:
        return {"status": "error", "message": str(exc)}

    if postgres_store_enabled():
        try:
            items, total, next_key = postgres_work_orders.load_page(
                limit=limit,
                cursor_created_at=decoded.created_at if decoded else "",
                cursor_id=decoded.record_id if decoded else "",
                role=actor_role(actor),
                user_id=actor_id(actor),
                line_scope=[str(value) for value in actor.get("line_scope", []) if isinstance(value, str)],
                status=status or "",
            )
        except ValueError as exc:
            return {"status": "error", "message": str(exc)}
        next_cursor = encode_cursor(*next_key) if next_key else ""
        return {
            "status": "ok",
            "orders": items,
            "total": total,
            "limit": limit,
            "next_cursor": next_cursor,
            "has_more": bool(next_cursor),
        }

    orders = _visible_orders(actor)
    if status:
        orders = [order for order in orders if order.get("status") == status]
    items, next_cursor, has_more = paginate_records(orders, limit=limit, cursor=decoded, id_field="id")
    return {
        "status": "ok",
        "orders": items,
        "total": len(orders),
        "limit": limit,
        "next_cursor": next_cursor,
        "has_more": has_more,
    }


@router.get(
    "/work-orders/stats",
    responses={200: {"model": WorkOrderStatsResponse}, **API_ERROR_RESPONSES},
)
async def api_order_stats(actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    return work_order_reporting.build_order_stats(
        _visible_orders(actor),
        statuses=STATUSES,
        priorities=PRIORITIES,
        knowledge_review_statuses=KB_REVIEW_STATUSES,
        status_labels=STATUS_LABELS,
        priority_labels=PRIORITY_LABELS,
    )


@router.get(
    "/work-orders/archive",
    responses={200: {"model": WorkOrderArchiveResponse}, **API_ERROR_RESPONSES},
)
async def api_work_order_archive(actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    archives, orders = _load_archived_orders()
    issues_by_id = _issue_map_by_id()
    return work_order_reporting.build_archive_response(
        archives,
        orders,
        actor=actor,
        issues_by_id=issues_by_id,
        can_view=can_view_work_order,
    )


@router.get(
    "/work-orders/{order_id}",
    responses={200: {"model": WorkOrderSuccessResponse}, **API_ERROR_RESPONSES},
)
async def api_get_order(order_id: str, actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    return work_order_queries.get_order_response(
        order_id,
        actor,
        _query_dependencies(),
    )


@router.get(
    "/work-orders/{order_id}/history",
    responses={200: {"model": WorkOrderHistoryResponse}, **API_ERROR_RESPONSES},
)
async def api_get_order_history(order_id: str, actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    return work_order_queries.get_history_response(
        order_id,
        actor,
        _query_dependencies(),
    )


@router.patch(
    "/work-orders/{order_id}",
    responses={200: {"model": WorkOrderMutationResponse}, **API_ERROR_RESPONSES},
)
@postgres_transactional
async def api_update_order(order_id: str, req: UpdateWorkOrder, actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    use_postgres = postgres_store_enabled()
    operation_dependencies = _operation_dependencies()
    loaded = work_order_operations.load_order(
        order_id,
        use_postgres=use_postgres,
        dependencies=operation_dependencies,
    )
    orders, idx, order = loaded.orders, loaded.index, loaded.order
    if order is None:
        return {"status": "error", "message": f"Work order {order_id} not found"}
    if order.get("deleted_at"):
        return {"status": "error", "message": f"Work order {order_id} is deleted"}
    linked_issue = None
    linked_issue_id = str(order.get("issue_id") or "")
    if linked_issue_id:
        linked_issue = operation_dependencies.get_issue(linked_issue_id)
    if not can_view_work_order(actor, order, linked_issue):
        return {"status": "error", "message": "Permission denied"}
    if not can_update_work_order(actor, order, req.status):
        return {"status": "error", "message": "Permission denied"}
    field_permission_error = _work_order_patch_permission_error(actor, req)
    if field_permission_error:
        return {"status": "error", "message": field_permission_error}
    current_version = work_order_mutations.normalized_version(order)
    if req.version is not None and req.version != current_version:
        return {"status": "error", "message": WORK_ORDER_STALE_UPDATE_MESSAGE}

    now = datetime.now().isoformat()
    linked_issue_snapshot = copy.deepcopy(linked_issue) if linked_issue is not None else None
    updated_by = actor_id(actor)
    mutation = work_order_mutations.apply_order_update(
        order,
        req,
        updated_by=updated_by,
        now=now,
        valid_statuses=STATUSES,
        stale_message=WORK_ORDER_STALE_UPDATE_MESSAGE,
        status_transition_error=_status_transition_error,
        closure_error=_closure_error,
        direct_verification_error=_direct_verification_error,
        refresh_knowledge_state=_refresh_knowledge_review_state,
        load_duplicate_orders=_load_orders,
        find_duplicate_order=_find_duplicate_knowledge_order,
        append_history=_append_order_history,
        calculate_field_changes=field_changes,
        increment_version=not use_postgres,
    )
    if mutation.error:
        return {"status": "error", "message": mutation.error}
    order = mutation.order
    before_order = mutation.before_order
    persisted = work_order_operations.persist_update_and_sync_issue(
        order,
        before_order,
        orders=orders,
        index=idx,
        use_postgres=use_postgres,
        linked_issue_snapshot=linked_issue_snapshot,
        dependencies=operation_dependencies,
    )
    if persisted.error:
        return {"status": "error", "message": persisted.error}
    order = persisted.order

    review_result = {
        "candidate": bool(order.get("kb_candidate")),
        "review_status": order.get("kb_review_status", "not_ready"),
    }
    return {
        "status": "ok",
        "order": order,
        "knowledge_review": review_result,
        "issue": persisted.synced_issue,
    }


@router.delete(
    "/work-orders/{order_id}",
    responses={200: {"model": WorkOrderDeleteResponse}, **API_ERROR_RESPONSES},
)
@postgres_transactional
async def api_delete_order(order_id: str, actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    if not is_admin(actor):
        return {"status": "error", "message": "Permission denied"}
    use_postgres = postgres_store_enabled()
    return work_order_operations.soft_delete_order(
        order_id,
        deleted_by=actor_id(actor),
        use_postgres=use_postgres,
        dependencies=_operation_dependencies(),
    )


# ----------------- KB feedback -----------------
async def _auto_feedback_to_kb(order: dict) -> dict:
    """Write an admin-approved work-order resolution to the knowledge base."""
    text = (
        f"[維修工單] Alarm: {order['alarm_code']}\n"
        f"機台: {order.get('machine_id', 'N/A')}\n"
        f"描述: {order.get('description', 'N/A')}\n"
        f"根本原因: {order.get('root_cause', 'N/A')}\n"
        f"實際維修動作: {order.get('repair_action', 'N/A')}\n"
        f"處理結果: {order['resolution']}\n"
        f"技師: {order.get('assigned_to', 'N/A')}\n"
        f"完成時間: {order.get('completed_at', 'N/A')}\n"
    )
    if order.get("notes"):
        text += f"備註: {order['notes']}\n"

    manual = order.get("manual", "808d")
    payload = {
        "text": text,
        "code": order.get("alarm_code", ""),
        "title": f"Work order {order['id']}",
        "page": 0,
        "source": "workorder",
    }

    try:
        from app_context import IngestTextRequest
        from app_context import get_engine
        from routes.ingest_routes import ingest_text_entry

        data = await ingest_text_entry(manual, IngestTextRequest(**payload))
        doc_id = data.get("doc_id")
        engine = get_engine(manual)
        indexed = bool(doc_id) and any(
            section.get("doc_id") == doc_id
            for section in engine.sections
        )
        validation_query = " ".join(
            str(order.get(field) or "")
            for field in ("description", "root_cause", "repair_action", "resolution")
        ).strip()
        retrieved = engine.retrieve(validation_query, top_k=5) if validation_query else []
        retrieval_hit = bool(doc_id) and any(
            result.get("meta", {}).get("doc_id") == doc_id
            for result in retrieved
        )
        ok = data.get("status") == "ok" and indexed
        return {
            "auto_ingested": ok,
            "collection": manual,
            "doc_id": doc_id,
            "text_preview": text[:200],
            "response": data,
            "validation": {
                "indexed": indexed,
                "retrieval_hit": retrieval_hit,
                "query": validation_query,
            },
        }
    except Exception as exc:
        return {
            "auto_ingested": False,
            "collection": manual,
            "text_preview": text[:200],
            "error": str(exc),
        }


@router.post(
    "/work-orders/{order_id}/knowledge-review",
    responses={
        **API_ERROR_RESPONSES,
        200: {"model": KnowledgeReviewResponse},
        400: {"model": KnowledgeReviewErrorResponse, "description": "Review or ingestion validation failed"},
    },
)
@postgres_transactional
async def review_work_order_knowledge(
    order_id: str,
    req: KnowledgeReviewRequest,
    actor: dict = Depends(get_actor),
):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    if not is_admin(actor):
        return {"status": "error", "message": "Permission denied"}

    use_postgres = postgres_store_enabled()
    orders = _load_orders()
    return await work_order_knowledge.review_order_knowledge(
        order_id,
        req,
        reviewer_id=actor_id(actor),
        orders=orders,
        use_postgres=use_postgres,
        dependencies=work_order_knowledge.ReviewDependencies(
            clean_text=_clean_text,
            refresh_review_state=_refresh_knowledge_review_state,
            candidate_ready=_knowledge_candidate_ready,
            find_duplicate=_find_duplicate_knowledge_order,
            append_history=_append_order_history,
            calculate_field_changes=field_changes,
            auto_feedback=_auto_feedback_to_kb,
            save_one=postgres_work_orders.save_one,
            save_all=_save_orders,
            stale_message=WORK_ORDER_STALE_UPDATE_MESSAGE,
        ),
    )


# ----------------- Excel import -----------------
# Expected Excel columns — order matters for positional fallback
EXCEL_FIELD_MAP = {
    "警報代碼": "alarm_code", "alarm_code": "alarm_code", "代碼": "alarm_code", "code": "alarm_code",
    "機台": "machine_id", "machine_id": "machine_id", "機台編號": "machine_id", "machine": "machine_id", "產線": "machine_id",
    "描述": "description", "description": "description", "問題描述": "description", "desc": "description",
    "指派": "assigned_to", "assigned_to": "assigned_to", "技師": "assigned_to", "assignee": "assigned_to",
    "處理結果": "resolution", "resolution": "resolution", "解決方案": "resolution",
    "優先": "priority", "priority": "priority", "優先級": "priority",
    "狀態": "status", "status": "status",
    "手冊": "manual", "manual": "manual",
    "來源": "source", "source": "source",
    "備註": "notes", "notes": "notes",
    "根本原因": "root_cause", "root_cause": "root_cause",
    "實際維修動作": "repair_action", "repair_action": "repair_action",
    "驗證人員": "verified_by", "verified_by": "verified_by",
}
POSITIONAL_FIELDS = [
    "alarm_code",
    "machine_id",
    "description",
    "assigned_to",
    "resolution",
    "priority",
    "status",
    "manual",
    "root_cause",
    "repair_action",
    "verified_by",
]


def _detect_columns(header_row: list) -> dict | None:
    """Try to map header cells to field names. Returns None if no match."""
    return detect_columns(header_row, EXCEL_FIELD_MAP)


def _validate_xlsx_archive(content: bytes) -> str:
    return validate_xlsx_archive(
        content,
        XlsxArchiveLimits(
            max_entries=XLSX_MAX_ENTRIES,
            max_uncompressed_bytes=XLSX_MAX_UNCOMPRESSED_BYTES,
            max_shared_strings_bytes=XLSX_MAX_SHARED_STRINGS_BYTES,
            max_compression_ratio=XLSX_MAX_COMPRESSION_RATIO,
        ),
    )


@router.post(
    "/work-orders/import-excel",
    responses={200: {"model": WorkOrderImportResponse}, **API_ERROR_RESPONSES},
)
async def import_excel(file: UploadFile = File(...), actor: dict = Depends(get_actor)):
    if not actor_id(actor):
        return {"status": "error", "message": "Not authenticated"}
    if not is_admin(actor):
        return {"status": "error", "message": "Permission denied"}
    """
    匯入 Excel 工單紀錄 (.xlsx)

    自動偵測 header:
    - 若第一列包含已知欄位名（如「警報代碼」「machine_id」），自動對應
    - 否則按預設順序: alarm_code, machine_id, description, assigned_to, resolution, priority, status, manual
    """
    safe_filename = os.path.basename(file.filename or "")
    if not safe_filename.lower().endswith(".xlsx"):
        return {"status": "error", "message": "Only .xlsx files are supported"}

    try:
        os.makedirs(DB_DIR, exist_ok=True)
        tmp_dir = tempfile.mkdtemp(dir=DB_DIR)
    except OSError as exc:
        return {"status": "error", "message": f"Unable to stage Excel import: {exc}"}
    tmp_path = os.path.join(tmp_dir, safe_filename)
    wb = None
    try:
        content = await file.read()
        if len(content) > EXCEL_UPLOAD_MAX_BYTES:
            max_mb = EXCEL_UPLOAD_MAX_BYTES / 1024 / 1024
            return {"status": "error", "message": f"Excel upload exceeds {max_mb:g} MB limit"}
        if not content.startswith(XLSX_MAGIC):
            return {"status": "error", "message": "Invalid XLSX file signature"}
        archive_error = _validate_xlsx_archive(content)
        if archive_error:
            return {"status": "error", "message": archive_error}
        with open(tmp_path, "wb") as f:
            f.write(content)

        from openpyxl import load_workbook
        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {"status": "error", "message": "Excel 檔案為空"}

        summary = import_workbook_rows(
            rows,
            field_map=EXCEL_FIELD_MAP,
            positional_fields=POSITIONAL_FIELDS,
            priorities=PRIORITIES,
            statuses=STATUSES,
            created_by=actor_id(actor),
            build_order=_build_order_dict,
            closure_error=_closure_error,
            refresh_knowledge_state=_refresh_knowledge_review_state,
            append_order_history=_append_order_history,
            calculate_field_changes=field_changes,
            persist_order=_persist_new_order,
        )

        return {
            "status": "ok",
            "filename": safe_filename,
            "imported": summary.imported,
            "skipped": summary.skipped,
            "errors": summary.errors[:10],
            "candidate_count": summary.candidate_count,
            "feedback_count": 0,
        }
    except Exception as e:
        return {"status": "error", "message": str(e)}
    finally:
        if wb is not None:
            try:
                wb.close()
            except Exception as exc:
                logger.warning("Failed to close imported workbook: %s", exc)
        shutil.rmtree(tmp_dir, ignore_errors=True)
