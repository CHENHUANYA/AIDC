import asyncio
import json
import os
from io import BytesIO
from pathlib import Path
from unittest.mock import Mock, patch

import pytest
from openpyxl import Workbook

import work_orders
from services import json_file_store


ADMIN = {"user_id": "admin01", "role": "admin", "line_scope": ["*"]}


class Upload:
    def __init__(self, filename: str, content: bytes):
        self.filename = filename
        self.content = content

    async def read(self) -> bytes:
        return self.content


def workbook_bytes(rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    for row in rows:
        worksheet.append(row)
    output = BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


@pytest.fixture
def local_order_store(tmp_path, monkeypatch):
    db_dir = tmp_path / "nested" / "alarm_db"
    monkeypatch.setattr(work_orders, "DB_DIR", str(db_dir))
    monkeypatch.setattr(work_orders, "WO_FILE", str(db_dir / "work_orders.json"))
    monkeypatch.setattr(work_orders, "ARCHIVE_DIR", str(db_dir / "archive"))
    monkeypatch.setattr(work_orders, "postgres_store_enabled", lambda: False)
    return db_dir


def test_excel_import_creates_missing_store_and_persists_each_valid_row_once(local_order_store):
    content = workbook_bytes(
        [
            ["alarm_code", "status", "root_cause", "repair_action", "resolution", "manual"],
            ["1000", "pending", "", "", "", "808d"],
            ["2000", "completed", "sensor failed", "replace sensor", "restored", "808d"],
            ["3000", "completed", "", "", "", "808d"],
        ]
    )

    result = asyncio.run(work_orders.import_excel(Upload("orders.xlsx", content), actor=ADMIN))

    assert result["status"] == "ok"
    assert result["imported"] == 2
    assert result["skipped"] == 1
    assert result["candidate_count"] == 1
    assert "requires root_cause and repair_action" in result["errors"][0]
    orders = work_orders._load_orders()
    assert {order["alarm_code"] for order in orders} == {"1000", "2000"}
    assert not any(order["alarm_code"] == "3000" for order in orders)
    completed = next(order for order in orders if order["alarm_code"] == "2000")
    assert completed["status"] == "completed"
    assert completed["kb_review_status"] == "pending_review"
    assert completed["work_order_history"][-1]["action"] == "import_status_override"
    assert completed["completed_by"] == "admin01"
    assert completed["work_order_history"][-1]["user_id"] == "admin01"
    assert not [path for path in local_order_store.iterdir() if path.is_dir()]


def test_invalid_completed_postgres_import_never_persists_pending_artifact(local_order_store, monkeypatch):
    content = workbook_bytes(
        [
            ["alarm_code", "status", "root_cause", "repair_action"],
            ["3000", "completed", "", ""],
        ]
    )
    monkeypatch.setattr(work_orders, "postgres_store_enabled", lambda: True)

    with patch.object(work_orders.postgres_work_orders, "save_one") as save_one:
        result = asyncio.run(work_orders.import_excel(Upload("orders.xlsx", content), actor=ADMIN))

    assert result["imported"] == 0
    assert result["skipped"] == 1
    save_one.assert_not_called()


def test_imported_verified_status_requires_post_import_verification(local_order_store):
    content = workbook_bytes([
        ["alarm_code", "status", "root_cause", "repair_action", "resolution", "verified_by"],
        ["2000", "verified", "sensor failed", "replaced sensor", "restored", "operator99"],
    ])

    result = asyncio.run(work_orders.import_excel(Upload("orders.xlsx", content), actor=ADMIN))
    order = work_orders._load_orders()[0]

    assert result["imported"] == 1
    assert order["status"] == "completed"
    assert order["verified_by"] == ""
    assert order["completed_by"] == "admin01"


def test_positional_import_reports_first_row_as_row_one(monkeypatch, local_order_store):
    content = workbook_bytes([["1000", "M-1", "symptom"]])
    monkeypatch.setattr(work_orders, "_persist_new_order", Mock(side_effect=OSError("persist failed")))

    result = asyncio.run(work_orders.import_excel(Upload("orders.xlsx", content), actor=ADMIN))

    assert result["imported"] == 0
    assert result["errors"] == ["Row 1: persist failed"]


def test_failed_candidate_persistence_does_not_increment_counts(monkeypatch, local_order_store):
    content = workbook_bytes(
        [
            ["alarm_code", "status", "root_cause", "repair_action", "resolution"],
            ["2000", "completed", "sensor failed", "replace sensor", "restored"],
        ]
    )
    monkeypatch.setattr(work_orders, "_persist_new_order", Mock(side_effect=OSError("persist failed")))

    result = asyncio.run(work_orders.import_excel(Upload("orders.xlsx", content), actor=ADMIN))

    assert result["imported"] == 0
    assert result["candidate_count"] == 0
    assert result["errors"] == ["Row 2: persist failed"]


def test_import_closes_workbook_after_row_iteration_failure(monkeypatch, local_order_store):
    content = workbook_bytes([["alarm_code"], ["1000"]])
    workbook = Mock()
    workbook.active.iter_rows.side_effect = RuntimeError("worksheet failed")

    with patch("openpyxl.load_workbook", return_value=workbook):
        result = asyncio.run(work_orders.import_excel(Upload("orders.xlsx", content), actor=ADMIN))

    assert result == {"status": "error", "message": "worksheet failed"}
    workbook.close.assert_called_once()
    assert not [path for path in local_order_store.iterdir() if path.is_dir()]


def test_import_rejects_legacy_xls_before_creating_storage(local_order_store):
    result = asyncio.run(work_orders.import_excel(Upload("orders.xls", b"legacy"), actor=ADMIN))

    assert result == {"status": "error", "message": "Only .xlsx files are supported"}
    assert not local_order_store.exists()


def test_atomic_work_order_save_preserves_original_on_replace_failure(local_order_store, monkeypatch):
    original = [{"id": "original"}]
    work_orders._save_orders(original)
    real_replace = json_file_store.os.replace

    def fail_destination(source, destination):
        if Path(destination) == Path(work_orders.WO_FILE):
            raise OSError("replace failed")
        real_replace(source, destination)

    monkeypatch.setattr(json_file_store.os, "replace", fail_destination)

    with pytest.raises(OSError, match="replace failed"):
        work_orders._save_orders([{"id": "replacement"}])

    assert json.loads(Path(work_orders.WO_FILE).read_text(encoding="utf-8")) == original
    assert not list(local_order_store.glob(".*.tmp"))


def test_atomic_json_serialization_failure_cleans_staged_file(local_order_store):
    with pytest.raises(TypeError):
        work_orders._save_orders([{"id": "bad", "unsupported": {"set"}}])

    assert not Path(work_orders.WO_FILE).exists()
    assert not list(local_order_store.glob(".*.tmp"))


def test_load_orders_rejects_wrong_root_and_skips_invalid_records(local_order_store):
    local_order_store.mkdir(parents=True)
    path = Path(work_orders.WO_FILE)
    path.write_text(json.dumps({"id": "not-a-list"}), encoding="utf-8")
    assert work_orders._load_orders() == []

    path.write_text(json.dumps(["bad", {"id": "valid", "version": "invalid"}]), encoding="utf-8")
    loaded = work_orders._load_orders()
    assert len(loaded) == 1
    assert loaded[0]["id"] == "valid"
    assert loaded[0]["version"] == 1
    assert loaded[0]["kb_review_status"] == "not_ready"

    path.write_text("{broken", encoding="utf-8")
    assert work_orders._load_orders() == []
    path.write_bytes(b"\xff\xfeinvalid")
    assert work_orders._load_orders() == []


def test_archive_loader_filters_invalid_records_and_survives_file_race(local_order_store, monkeypatch):
    archive_dir = Path(work_orders.ARCHIVE_DIR)
    archive_dir.mkdir(parents=True)
    valid = archive_dir / "work_orders_archive_20260818.json"
    corrupt = archive_dir / "work_orders_archive_20260817.json"
    gone = archive_dir / "work_orders_archive_20260816.json"
    ignored = archive_dir / "notes.json"
    valid.write_text(json.dumps([{"id": "archived"}, "invalid"]), encoding="utf-8")
    corrupt.write_text("{broken", encoding="utf-8")
    gone.write_text("[]", encoding="utf-8")
    ignored.write_text("[]", encoding="utf-8")
    real_getmtime = os.path.getmtime

    def getmtime(path):
        if Path(path) == gone:
            raise FileNotFoundError(path)
        return real_getmtime(path)

    monkeypatch.setattr(work_orders.os.path, "getmtime", getmtime)

    archives, orders = work_orders._load_archived_orders()

    assert [archive["file"] for archive in archives] == [valid.name, corrupt.name]
    assert [archive["count"] for archive in archives] == [1, 0]
    assert orders == [{"id": "archived", "archive_file": valid.name}]


def test_archive_loader_handles_directory_listing_failure(local_order_store, monkeypatch):
    Path(work_orders.ARCHIVE_DIR).mkdir(parents=True)
    monkeypatch.setattr(work_orders.os, "listdir", Mock(side_effect=PermissionError("denied")))

    assert work_orders._load_archived_orders() == ([], [])


def test_archive_api_filters_visibility_and_guards_auth(monkeypatch):
    archived = [
        {"id": "visible", "completed_at": "2026-08-18", "issue_id": "I-1"},
        {"id": "hidden", "completed_at": "2026-08-19", "issue_id": "I-2"},
    ]
    monkeypatch.setattr(work_orders, "_load_archived_orders", lambda: ([{"file": "archive.json"}], archived))
    monkeypatch.setattr(work_orders, "_issue_map_by_id", lambda: {"I-1": {}, "I-2": {}})
    monkeypatch.setattr(work_orders, "can_view_work_order", lambda _actor, order, _issue: order["id"] == "visible")

    assert asyncio.run(work_orders.api_work_order_archive(actor={}))["message"] == "Not authenticated"
    result = asyncio.run(work_orders.api_work_order_archive(actor=ADMIN))

    assert result["total"] == 1
    assert result["orders"][0]["id"] == "visible"
